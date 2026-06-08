"""文档导入模块 - 支持 .docx / .txt 文件解析与句子切割 + 中英文过滤"""

import io
import re
from docx import Document


# 句子分隔符：中英文句号、问号、感叹号
_SENTENCE_SPLITTER = re.compile(r"(?<=[。！？.!?])(?:\s+|\n|)(?=[^\s])")


def parse_uploaded_file(uploaded_file) -> list[dict]:
    """
    解析上传的文件，按句子切分，生成 Sentence ID。

    Args:
        uploaded_file: Streamlit UploadedFile 对象

    Returns:
        [
            {"sentence_id": "S1", "source_text": "句子内容..."},
            {"sentence_id": "S2", "source_text": "句子内容..."},
            ...
        ]

    Raises:
        ValueError: 不支持的文件格式
    """
    filename = uploaded_file.name.lower()

    if filename.endswith(".txt"):
        content = uploaded_file.read().decode("utf-8")
        raw_paragraphs = _split_paragraphs(content)
    elif filename.endswith(".docx"):
        content_bytes = uploaded_file.read()
        raw_paragraphs = _parse_docx(content_bytes)
    else:
        raise ValueError(f"不支持的文件格式：{uploaded_file.name}，请上传 .docx 或 .txt 文件")

    # 将每个段落进一步切分为句子
    sentences = _split_sentences(raw_paragraphs)

    # 过滤空句子，生成 Sentence ID
    results = []
    idx = 0
    for sent in sentences:
        text = sent.strip()
        if not text:
            continue
        idx += 1
        results.append({
            "sentence_id": f"S{idx}",
            "source_text": text,
        })

    return results


def _split_paragraphs(content: str) -> list[str]:
    """将纯文本按自然段落切分（先按空行，再按单换行）"""
    content = content.replace("\r\n", "\n").replace("\r", "\n")

    if "\n\n" in content:
        parts = content.split("\n\n")
    else:
        parts = content.split("\n")

    return [p.strip() for p in parts if p.strip()]


def _parse_docx(content_bytes: bytes) -> list[str]:
    """从 .docx 文件中提取段落文本，每个 <w:p> 元素作为一个段落"""
    doc = Document(io.BytesIO(content_bytes))
    paragraphs = []
    for para in doc.paragraphs:
        text = para.text
        if text is not None and text.strip():
            paragraphs.append(text.strip())
    return paragraphs


def _split_sentences(paragraphs: list[str]) -> list[str]:
    """将段落列表进一步按句子切分"""
    sentences = []
    for para in paragraphs:
        # 用正则按句子分隔符切分
        parts = _SENTENCE_SPLITTER.split(para)
        for part in parts:
            s = part.strip()
            if s:
                sentences.append(s)
    return sentences


# ═══════════════════════════════════════════════════════════════
#  中英文检测与过滤
# ═══════════════════════════════════════════════════════════════

_CHINESE_PATTERN = re.compile(r"[一-鿿㐀-䶿豈-﫿]")


def has_chinese(text: str) -> bool:
    """判断文本中是否包含中文字符"""
    return bool(_CHINESE_PATTERN.search(text))


def filter_chinese_only(segments: list[dict]) -> list[dict]:
    """只保留包含中文的句子，保留原始 Sentence ID 不变"""
    return [seg for seg in segments if has_chinese(seg["source_text"])]


def filter_english_only(segments: list[dict]) -> list[dict]:
    """只保留不包含中文的句子（纯英文）"""
    return [seg for seg in segments if not has_chinese(seg["source_text"])]


# ═══════════════════════════════════════════════════════════════
#  Excel 导出（自动换行 + 自适应列宽）
# ═══════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def export_to_xlsx(segments: list[dict], col_label: str = "Source Text") -> bytes:
    """
    将句子列表导出为 .xlsx 文件（bytes），支持：
    - 文本自动换行
    - 自适应列宽（对长句友好）
    - 冻结表头行

    Args:
        segments: 句子列表 [{"sentence_id": "S1", "source_text": "..."}, ...]
        col_label: 第二列的列名

    Returns:
        Excel 文件的 bytes，可直接用于 st.download_button
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sentences"

    # 表头
    headers = ["Sentence ID", col_label]
    ws.append(headers)

    # 表头样式
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据行
    for seg in segments:
        ws.append([seg["sentence_id"], seg["source_text"]])

    # ── 列宽与换行 ──────────────────────────────────
    # Sentence ID 列：固定宽度
    ws.column_dimensions["A"].width = 14

    # Source Text 列：根据内容自适应，上限 80
    max_len = 0
    for seg in segments:
        # 取最长行（按换行符拆分后最长的单行）
        for line in seg["source_text"].split("\n"):
            # 粗略估计：中文字符占 2，ASCII 占 1
            line_len = sum(2 if ord(c) > 127 else 1 for c in line)
            if line_len > max_len:
                max_len = line_len

    col_width = min(max_len + 4, 80)  # 上限 80 字符宽，保底 40
    col_width = max(col_width, 40)
    ws.column_dimensions[get_column_letter(2)].width = col_width

    # 所有数据行设置自动换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 冻结表头
    ws.freeze_panes = "A2"

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
