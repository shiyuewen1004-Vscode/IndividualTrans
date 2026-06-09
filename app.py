"""翻译记忆学习系统 - Streamlit 主应用"""

import io
import csv
import streamlit as st
from translator import translate, PROVIDER_LABELS, DIRECTION_LABELS
from config import DEFAULT_PROVIDER
from database import init_db, get_prompt_override, save_prompt_override, delete_prompt_override, get_all_prompt_overrides
from tracker import record_modification, confirm_rule, ignore_rule, defer_rule
from document import parse_uploaded_file, filter_chinese_only, filter_english_only
from term_annotator import (
    load_terminology, get_all_domains, get_term_count, get_csv_path,
    detect_domains, detect_domains_by_terms, annotate_text,
    get_unique_terms, append_to_csv, DOMAIN_KEYWORDS,
)
from prompt_center import (
    STYLE_LABELS, STYLE_CONFIGS,
    get_default_prompt_text, get_effective_principles,
)

st.set_page_config(page_title="翻译记忆学习系统", page_icon="🌐", layout="wide")
init_db()

# ── Session State ────────────────────────────────────
if "style_prompt" not in st.session_state:
    st.session_state.style_prompt = None
if "provider" not in st.session_state:
    st.session_state.provider = DEFAULT_PROVIDER
if "segments" not in st.session_state:
    st.session_state.segments = []
if "doc_translations" not in st.session_state:
    st.session_state.doc_translations = {}
if "doc_direction" not in st.session_state:
    st.session_state.doc_direction = "zh2en"
if "doc_edits" not in st.session_state:
    st.session_state.doc_edits = {}
if "doc_editing" not in st.session_state:
    st.session_state.doc_editing = False
if "app_domain" not in st.session_state:
    st.session_state.app_domain = "其他"
if "prompt_template" not in st.session_state:
    st.session_state.prompt_template = "default"
if "custom_prompt_text" not in st.session_state:
    st.session_state.custom_prompt_text = ""
if "last_retrieval" not in st.session_state:
    st.session_state.last_retrieval = None
# ── 术语标注 Session State ────────────────────────────
if "term_input_text" not in st.session_state:
    st.session_state.term_input_text = ""
if "term_annotated_segments" not in st.session_state:
    st.session_state.term_annotated_segments = []
if "term_matched_csv" not in st.session_state:
    st.session_state.term_matched_csv = []
if "term_matched_user" not in st.session_state:
    st.session_state.term_matched_user = []
if "term_user_added" not in st.session_state:
    st.session_state.term_user_added = []
if "term_domain_label" not in st.session_state:
    st.session_state.term_domain_label = ""
if "term_best_domain" not in st.session_state:
    st.session_state.term_best_domain = "全部领域"
if "term_annotated" not in st.session_state:
    st.session_state.term_annotated = False
if "term_adding_mode" not in st.session_state:
    st.session_state.term_adding_mode = False

# ── 侧边栏 ────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 模型")
    provider = st.selectbox(
        "翻译引擎",
        options=list(PROVIDER_LABELS.keys()),
        format_func=lambda k: PROVIDER_LABELS[k],
        key="provider_selector",
    )
    st.session_state.provider = provider

    st.divider()
    from prompt_center import PROMPT_LABELS
    st.session_state.prompt_template = st.selectbox(
        "Prompt",
        options=list(PROMPT_LABELS.keys()),
        format_func=lambda k: PROMPT_LABELS[k],
        key="prompt_sidebar",
    )
    if st.session_state.prompt_template == "custom":
        st.session_state.custom_prompt_text = st.text_area(
            "自定义", value=st.session_state.custom_prompt_text or "",
            height=100, label_visibility="collapsed",
        )

# ── 主界面 ────────────────────────────────────────────
st.title("🌐 翻译记忆学习系统")

tab1, tab2, tab3, tab4, tab5 = st.tabs(["📄 文档翻译", "📁 我的文件", "📚 记忆库", "🎨 Prompt", "🔍 术语标注"])

# ═══════════════════════════════════════════════════════
#  Tab 1：文档翻译
# ═══════════════════════════════════════════════════════
with tab1:
    uploaded_file = st.file_uploader(
        "上传文档", type=["docx", "txt"],
        help="支持 .docx 和 .txt 格式", key="doc_uploader",
    )

    if uploaded_file is not None:
        if st.button("🔍 解析文档", type="primary"):
            with st.spinner("正在解析文档..."):
                try:
                    st.session_state.segments = parse_uploaded_file(uploaded_file)
                    from database import add_file_history
                    add_file_history(uploaded_file.name, len(st.session_state.segments))
                    st.success(f"解析完成！共 {len(st.session_state.segments)} 个句子")
                except Exception as e:
                    st.error(f"解析失败：{e}")
                    st.session_state.segments = []

    if st.session_state.segments:
        lang_filter = st.radio(
            "📋 句子过滤", options=["全部显示", "仅显示中文", "仅显示英文"],
            horizontal=True, key="lang_filter",
        )
        if lang_filter == "仅显示中文":
            display_segments = filter_chinese_only(st.session_state.segments)
            col_label = "Source Text（中文）"
        elif lang_filter == "仅显示英文":
            display_segments = filter_english_only(st.session_state.segments)
            col_label = "Source Text（英文）"
        else:
            display_segments = st.session_state.segments
            col_label = "Source Text"

        if display_segments:
            st.markdown(f"**共 {len(display_segments)} 个句子**（全部 {len(st.session_state.segments)} 个）")

            col_dir, col_dom, col_btn = st.columns([1, 1, 2])
            with col_dir:
                st.session_state.doc_direction = st.selectbox(
                    "翻译方向", options=list(DIRECTION_LABELS.keys()),
                    format_func=lambda k: DIRECTION_LABELS[k], key="doc_direction_selector",
                )
            with col_dom:
                st.session_state.app_domain = st.selectbox(
                    "Domain", options=["经济金融", "传统文化", "政治外交", "化学化工", "教育", "法律", "医学", "其他"], key="doc_domain",
                )
            with col_btn:
                st.write(""); st.write("")
                translate_all_clicked = st.button("🌐 翻译全部", type="primary", use_container_width=True)

            if translate_all_clicked:
                st.session_state.doc_translations = {}
                st.session_state.last_retrieval = None
                progress_bar = st.progress(0)
                total = len(display_segments)
                for i, seg in enumerate(display_segments):
                    try:
                        translation, retrieval = translate(
                            seg["source_text"], provider=st.session_state.provider,
                            direction=st.session_state.doc_direction,
                            prompt_template=st.session_state.prompt_template,
                            custom_prompt=st.session_state.custom_prompt_text,
                        )
                        st.session_state.doc_translations[seg["sentence_id"]] = translation
                        if retrieval:
                            st.session_state.last_retrieval = retrieval
                    except Exception as e:
                        st.session_state.doc_translations[seg["sentence_id"]] = f"❌ {e}"
                    progress_bar.progress((i + 1) / total)
                st.success(f"翻译完成！共 {total} 句")

            # ── 检索结果展示 ────────────────────────────
            if st.session_state.get("last_retrieval"):
                r = st.session_state.last_retrieval
                with st.expander(
                    f"🔍 本次翻译使用 · 模型：{PROVIDER_LABELS[st.session_state.provider]} · "
                    f"Prompt：{st.session_state.prompt_template} · "
                    f"记忆库命中：{r['count']}",
                    expanded=False,
                ):
                    for h in r["hits"]:
                        st.caption(f"{h['source_text']} → {h['target_text']}")

            # 修改模式切换
            if st.session_state.doc_translations:
                col_tbl, col_edit, _ = st.columns([2, 1, 1])
                with col_edit:
                    if not st.session_state.doc_editing:
                        if st.button("✏️ 批量修改", use_container_width=True):
                            st.session_state.doc_editing = True
                            for seg in display_segments:
                                sid = seg["sentence_id"]
                                if sid not in st.session_state.doc_edits:
                                    st.session_state.doc_edits[sid] = st.session_state.doc_translations.get(sid, "")
                            st.rerun()
                    else:
                        if st.button("👁 查看模式", use_container_width=True):
                            st.session_state.doc_editing = False
                            st.rerun()

            # 编辑模式
            if st.session_state.doc_editing and st.session_state.doc_translations:
                st.info("📝 横向对比修改每条译文，可逐句保存或一键保存全部")
                for seg in display_segments:
                    sid = seg["sentence_id"]
                    ai_text = st.session_state.doc_translations.get(sid, "")
                    saved_mark = " ✅" if f"saved_{sid}" in st.session_state and st.session_state[f"saved_{sid}"] else ""
                    with st.expander(f"{sid}{saved_mark} — {seg['source_text'][:80]}...", expanded=False):
                        col_src, col_ai, col_edit = st.columns([1, 1, 1])
                        with col_src:
                            st.markdown("**原文**")
                            st.markdown(f'<div style="background:#f0f2f6;padding:12px;border-radius:8px;min-height:120px;white-space:pre-wrap;font-size:14px;">{seg["source_text"]}</div>', unsafe_allow_html=True)
                        with col_ai:
                            st.markdown("**AI Translation**")
                            st.markdown(f'<div style="background:#e8f5e9;padding:12px;border-radius:8px;min-height:120px;white-space:pre-wrap;font-size:14px;word-wrap:break-word;">{ai_text}</div>', unsafe_allow_html=True)
                        with col_edit:
                            st.markdown("**Human Translation**")
                            edited = st.text_area("Human Translation", value=st.session_state.doc_edits.get(sid, ai_text), height=140, label_visibility="collapsed", key=f"edit_{sid}")
                            st.session_state.doc_edits[sid] = edited
                            if st.button(f"💾 保存 {sid}", key=f"save_{sid}", use_container_width=True):
                                original, modified = ai_text, edited
                                if original.strip() != modified.strip():
                                    result = record_modification(source=seg["source_text"], original=original, modified=modified)
                                    if result["should_prompt"]:
                                        st.session_state.style_prompt = {"rule_id": result["rule_id"], "original": result["original_phrase"], "modified": result["modified_phrase"], "count": result["count"], "status": result["status"]}
                                    from database import insert_asset
                                    insert_asset(source=seg["source_text"], target=modified, domain=st.session_state.app_domain)
                                st.session_state[f"saved_{sid}"] = True
                                st.success(f"{sid} 已保存 → Memory Base")
                                st.rerun()

                col_save_all, _ = st.columns([1, 3])
                with col_save_all:
                    if st.button("💾 保存全部修改", type="primary", use_container_width=True):
                        saved_count, last_prompt = 0, None
                        from database import insert_asset
                        for seg in display_segments:
                            sid = seg["sentence_id"]
                            original = st.session_state.doc_translations.get(sid, "")
                            modified = st.session_state.doc_edits.get(sid, "")
                            if original.strip() != modified.strip():
                                result = record_modification(source=seg["source_text"], original=original, modified=modified)
                                if result["should_prompt"]:
                                    last_prompt = {"rule_id": result["rule_id"], "original": result["original_phrase"], "modified": result["modified_phrase"], "count": result["count"], "status": result["status"]}
                                insert_asset(source=seg["source_text"], target=modified, domain=st.session_state.app_domain)
                            st.session_state[f"saved_{sid}"] = True
                            saved_count += 1
                        st.session_state.doc_editing = False
                        if last_prompt:
                            st.session_state.style_prompt = last_prompt
                        st.success(f"已保存 {saved_count} 条 → Memory Base")
                        st.rerun()

            # 表格
            if st.session_state.doc_translations:
                if st.session_state.doc_edits:
                    headers = ["Sentence ID", "Source Text", "AI Translation", "Human Translation"]
                    rows_html = "".join(
                        f"<tr><td style='white-space:nowrap;vertical-align:top;padding:8px;'>{seg['sentence_id']}</td>"
                        f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{seg['source_text']}</td>"
                        f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{st.session_state.doc_translations.get(seg['sentence_id'], '')}</td>"
                        f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{st.session_state.doc_edits.get(seg['sentence_id'], '')}</td></tr>"
                        for seg in display_segments
                    )
                else:
                    headers = ["Sentence ID", "Source Text", "AI Translation"]
                    rows_html = "".join(
                        f"<tr><td style='white-space:nowrap;vertical-align:top;padding:8px;'>{seg['sentence_id']}</td>"
                        f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{seg['source_text']}</td>"
                        f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{st.session_state.doc_translations.get(seg['sentence_id'], '')}</td></tr>"
                        for seg in display_segments
                    )
            else:
                headers = ["Sentence ID", "Source Text"]
                rows_html = "".join(
                    f"<tr><td style='white-space:nowrap;vertical-align:top;padding:8px;'>{seg['sentence_id']}</td>"
                    f"<td style='white-space:pre-wrap;word-wrap:break-word;vertical-align:top;padding:8px;'>{seg['source_text']}</td></tr>"
                    for seg in display_segments
                )

            header_html = "".join(f"<th style='position:sticky;top:0;background:#e0e0e0;padding:10px;text-align:left;'>{h}</th>" for h in headers)
            st.markdown(f"<div style='max-height:500px;overflow-y:auto;border:1px solid #ddd;border-radius:8px;'><table style='width:100%;border-collapse:collapse;font-size:14px;'><thead><tr>{header_html}</tr></thead><tbody>{rows_html}</tbody></table></div>", unsafe_allow_html=True)
            st.caption(f"共 {len(display_segments)} 条")

            # CSV 导出
            csv_buffer = io.StringIO(newline="")
            writer = csv.writer(csv_buffer, quoting=csv.QUOTE_ALL)
            if st.session_state.doc_translations:
                if st.session_state.doc_edits:
                    writer.writerow(["Sentence ID", "Source Text", "AI Translation", "Human Translation"])
                    for seg in display_segments:
                        sid = seg["sentence_id"]
                        writer.writerow([sid, seg["source_text"], st.session_state.doc_translations.get(sid, ""), st.session_state.doc_edits.get(sid, "")])
                else:
                    writer.writerow(["Sentence ID", "Source Text", "AI Translation"])
                    for seg in display_segments:
                        writer.writerow([seg["sentence_id"], seg["source_text"], st.session_state.doc_translations.get(seg["sentence_id"], "")])
            else:
                writer.writerow(["Sentence ID", "Source Text"])
                for seg in display_segments:
                    writer.writerow([seg["sentence_id"], seg["source_text"]])
            csv_bytes = csv_buffer.getvalue().encode("utf-8-sig")
            import base64
            st.markdown(f'<a href="data:text/csv;charset=utf-8;base64,{base64.b64encode(csv_bytes).decode()}" download="sentences.csv" style="display:inline-block;padding:6px 16px;background:#4CAF50;color:#fff;text-decoration:none;border-radius:6px;font-size:14px;">📥 下载 CSV</a>', unsafe_allow_html=True)
        else:
            st.warning("过滤后无匹配句子")
    else:
        st.info("👆 请上传 .docx 或 .txt 文件，然后点击「解析文档」")

# ═══════════════════════════════════════════════════════
#  Tab 2：我的文件
# ═══════════════════════════════════════════════════════
with tab2:
    st.subheader("📁 我的文件")
    from database import get_file_history, delete_file_history

    files = get_file_history(limit=50)
    if files:
        for f in files:
            col_f, col_d = st.columns([6, 1])
            with col_f:
                st.markdown(f"📄 **{f['filename']}** — {f['sentence_count']} 句 — _{f['uploaded_at']}_")
            with col_d:
                if st.button("🗑", key=f"tab2_del_{f['id']}"):
                    delete_file_history(f["id"])
                    st.rerun()
    else:
        st.info("暂无上传记录")

# ═══════════════════════════════════════════════════════
#  Tab 3：记忆库
# ═══════════════════════════════════════════════════════
with tab3:
    from database import get_all_assets, get_asset_stats, delete_asset, insert_asset

    stats = get_asset_stats()
    m1, m2 = st.columns(2)
    with m1: st.metric("📊 Total Assets", stats["total"])
    with m2: st.metric("✅ Active", stats.get("active", stats["total"]))

    # Import
    with st.expander("📥 Import Excel / CSV", expanded=False):
        imp_file = st.file_uploader("上传", type=["xlsx", "csv"], key="mb_import_file", label_visibility="collapsed")
        if imp_file and st.button("📥 开始导入", type="primary", use_container_width=True):
            try:
                fn = imp_file.name.lower()
                if fn.endswith(".csv"):
                    rows = list(csv.DictReader(io.StringIO(imp_file.read().decode("utf-8-sig"))))
                elif fn.endswith(".xlsx"):
                    from openpyxl import load_workbook
                    wb = load_workbook(imp_file, read_only=True); ws = wb.active
                    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
                    wb.close()
                else:
                    st.error("不支持的文件格式"); rows = []
                if rows:
                    missing = {"source_text", "target_text"} - set(rows[0].keys())
                    if missing:
                        st.error(f"❌ 缺少必须字段：{', '.join(missing)}")
                    else:
                        imported = skipped = 0
                        for r in rows:
                            src = (r.get("source_text") or "").strip()
                            tgt = (r.get("target_text") or "").strip()
                            if not src or not tgt: skipped += 1; continue
                            insert_asset(source=src, target=tgt, domain=(r.get("domain") or "其他").strip())
                            imported += 1
                        st.success(f"✅ 导入 {imported} 条" + (f"，跳过 {skipped} 条" if skipped else "")); st.rerun()
            except Exception as e:
                st.error(f"导入失败：{e}")

    # Search + filters
    c1, c2, c3 = st.columns(3)
    with c1: keyword = st.text_input("🔍 搜索", key="mb_search")
    with c2: domain_filter = st.selectbox("Domain", options=["全部", "经济金融", "传统文化", "政治外交", "化学化工", "教育", "法律", "医学", "其他"], key="mb_domain")
    with c3: status_filter = st.selectbox("Status", options=["全部", "active", "draft", "archived"], key="mb_status")

    assets = get_all_assets(
        domain=None if domain_filter == "全部" else domain_filter,
        status=None if status_filter == "全部" else status_filter,
        keyword=keyword.strip() or None,
    )

    selected_indices = []
    if "mb_selected_rows" in st.session_state:
        sel = st.session_state.mb_selected_rows
        if isinstance(sel, dict) and "selection" in sel:
            selected_indices = sel["selection"].get("rows", [])
        elif hasattr(sel, "selection") and hasattr(sel.selection, "rows"):
            selected_indices = sel.selection.rows

    ci, cd = st.columns([4, 1])
    with ci: st.markdown(f"**共 {len(assets)} 条**（总计 {stats['total']} 条）")
    with cd:
        if assets and selected_indices:
            if st.button(f"🗑 删除 ({len(selected_indices)})", type="secondary", use_container_width=True):
                for idx in selected_indices: delete_asset(assets[idx]["id"])
                st.session_state.mb_selected_rows = {}
                st.success(f"已删除 {len(selected_indices)} 条"); st.rerun()

    if not assets:
        st.info("🙅 No Assets Found.")
    else:
        import pandas as pd
        df = pd.DataFrame([{"ID": a["id"], "Source Text": a["source_text"], "Target Text": a["target_text"], "Domain": a["domain"], "Status": a["status"], "Updated Time": a["updated_time"]} for a in assets])
        st.dataframe(df, use_container_width=True, height=550, hide_index=True, selection_mode="multi-row", on_select="rerun", key="mb_selected_rows")


# ═══════════════════════════════════════════════════════════════
#  Tab 4：术语标注
# ═══════════════════════════════════════════════════════════════
with tab4:
    st.subheader("🔍 术语自动标注")

    # ── 术语库概况 ──
    csv_path = get_csv_path()
    terminology = load_terminology(csv_path)
    all_term_domains = get_all_domains(csv_path)
    total_terms = get_term_count(csv_path)

    with st.expander("📊 术语库概况", expanded=False):
        col_a, col_b = st.columns(2)
        with col_a:
            st.metric("领域数", len(all_term_domains))
        with col_b:
            st.metric("术语总数", total_terms)
        if all_term_domains:
            domain_lines = ""
            for d in all_term_domains:
                cnt = len(terminology.get(d, []))
                domain_lines += f"- **{d}**：{cnt} 条\n"
            st.markdown(domain_lines)

    # ── 图例 ──
    st.caption("📋 图例：🔵 **青色高亮** = 术语库匹配 | 🔴 **红色高亮** = 用户补充")

    # ── 文本输入区 ──
    st.markdown("### 📝 输入文本")
    user_input = st.text_area(
        "请输入需要标注的文本（中英文均可）",
        value=st.session_state.term_input_text,
        height=180,
        key="term_text_input",
        placeholder="在此粘贴或输入文本...\n例如：患者出现发热和头痛症状，医生建议进行影像检查。The patient underwent surgery yesterday.",
    )

    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])

    with col_btn1:
        annotate_clicked = st.button("🔍 标注术语", type="primary", use_container_width=True)

    with col_btn2:
        clear_clicked = st.button("🔄 清空", use_container_width=True)

    if clear_clicked:
        st.session_state.term_input_text = ""
        st.session_state.term_annotated_segments = []
        st.session_state.term_matched_csv = []
        st.session_state.term_matched_user = []
        st.session_state.term_user_added = []
        st.session_state.term_domain_label = ""
        st.session_state.term_best_domain = "全部领域"
        st.session_state.term_annotated = False
        st.session_state.term_adding_mode = False
        st.rerun()

    # ── 标注逻辑 ──
    if annotate_clicked and user_input.strip():
        st.session_state.term_input_text = user_input
        text = user_input.strip()

        # 1. 领域检测（关键词 + 术语回退）
        detected = detect_domains(text)
        if not detected:
            detected = detect_domains_by_terms(text, terminology)

        # 2. 确定活跃术语
        if not detected:
            active_terms: list[tuple[str, str]] = []
            for t_list in terminology.values():
                active_terms.extend(t_list)
            domain_label = "全部领域"
            best_domain = "全部领域"
        else:
            active_domains = [d for d, _ in detected]
            active_terms = []
            for d in active_domains:
                active_terms.extend(terminology.get(d, []))
            best_domain = detected[0][0]
            if len(detected) == 1:
                domain_label = best_domain
            else:
                domain_label = "、".join(d for d, _ in detected)

        # 3. 合并用户已添加的术语
        merged_terms = list(active_terms)
        for ch, en in st.session_state.term_user_added:
            if (ch, en) not in merged_terms:
                merged_terms.append((ch, en))

        # 4. 标注
        result = annotate_text(text, merged_terms)
        st.session_state.term_annotated_segments = result["annotated_segments"]
        st.session_state.term_matched_csv = result["matched_csv"]
        st.session_state.term_matched_user = result["matched_user"]
        st.session_state.term_domain_label = domain_label
        st.session_state.term_best_domain = best_domain
        st.session_state.term_annotated = True
        st.session_state.term_adding_mode = False

    # ── 标注结果展示 ──
    if st.session_state.term_annotated and st.session_state.term_annotated_segments:
        segments = st.session_state.term_annotated_segments
        matched_csv = st.session_state.term_matched_csv
        matched_user = st.session_state.term_matched_user
        domain_label = st.session_state.term_domain_label

        total_matches = len(matched_csv) + len(matched_user)
        unique = get_unique_terms(matched_csv, matched_user)

        st.divider()

        # ── 领域标签 ──
        detected = detect_domains(st.session_state.term_input_text)
        if not detected:
            detected = detect_domains_by_terms(st.session_state.term_input_text, terminology)
        if detected:
            tags_html = " ".join(
                f'<span style="display:inline-block;background:#7c3aed;color:#fff;padding:2px 10px;border-radius:12px;font-size:13px;margin:2px;">{d} ({s}词)</span>'
                for d, s in detected
            )
            st.markdown(f"**🏷️ 检测领域：** {tags_html}", unsafe_allow_html=True)
        else:
            st.info("🏷️ 检测领域：全部领域（未匹配到特定领域关键词）")

        st.markdown(f"**📊 匹配统计：{total_matches} 处 / {len(unique)} 个术语** | 领域：{domain_label}")

        # ── 高亮渲染文本 ──
        st.markdown("### 📋 标注文本")
        html_parts = ['<div style="background:#1e1e1e;padding:20px;border-radius:10px;line-height:2;font-size:15px;white-space:pre-wrap;word-wrap:break-word;font-family:-apple-system,BlinkMacSystemFont,sans-serif;">']
        for seg in segments:
            if seg["is_match"]:
                if seg.get("is_user"):
                    # 红色：用户补充
                    html_parts.append(
                        f'<span style="background:#dc3545;color:#fff;padding:2px 6px;border-radius:4px;font-weight:bold;" '
                        f'title="中文：{seg["chinese"]} | 英文：{seg["english"]}（用户补充）">'
                        f'{seg["text"]}</span>'
                    )
                else:
                    # 青色：术语库匹配
                    html_parts.append(
                        f'<span style="background:#17a2b8;color:#fff;padding:2px 6px;border-radius:4px;font-weight:bold;" '
                        f'title="中文：{seg["chinese"]} | 英文：{seg["english"]}">'
                        f'{seg["text"]}</span>'
                    )
            else:
                html_parts.append(
                    f'<span style="color:#e0e0e0;">{seg["text"]}</span>'
                )
        html_parts.append('</div>')
        st.markdown("".join(html_parts), unsafe_allow_html=True)

        # ── 术语对照表 ──
        if unique:
            st.markdown("### 📖 术语对照表")
            import pandas as pd
            table_data = []
            for ch, en, is_user in unique:
                source = "👤 用户补充" if is_user else "📚 术语库"
                table_data.append({"中文术语": ch, "英文术语": en, "来源": source})
            st.dataframe(
                pd.DataFrame(table_data),
                use_container_width=True,
                hide_index=True,
            )

        # ── CSV 导出 ──
        if unique:
            csv_buffer = io.StringIO(newline="")
            writer = csv.writer(csv_buffer, quoting=csv.QUOTE_ALL)
            writer.writerow(["中文术语", "英文术语", "领域"])
            for ch, en, _ in unique:
                writer.writerow([ch, en, st.session_state.term_best_domain])
            csv_bytes = csv_buffer.getvalue().encode("utf-8-sig")
            import base64
            st.markdown(
                f'<a href="data:text/csv;charset=utf-8;base64,{base64.b64encode(csv_bytes).decode()}" '
                f'download="terminology_export.csv" '
                f'style="display:inline-block;padding:6px 16px;background:#17a2b8;color:#fff;text-decoration:none;border-radius:6px;font-size:14px;">'
                f'📥 下载对照表 CSV</a>',
                unsafe_allow_html=True,
            )

        # ── 补充遗漏术语 ──
        st.divider()

        if not st.session_state.term_adding_mode:
            st.markdown("### ➕ 补充遗漏术语")
            st.caption("发现有术语未被标注？点击下方按钮手动补充，新增术语将自动写入术语库 CSV。")
            if st.button("➕ 添加遗漏术语", type="secondary"):
                st.session_state.term_adding_mode = True
                st.rerun()

        if st.session_state.term_adding_mode:
            st.markdown("### ✏️ 添加遗漏术语")
            st.caption("格式：**中文术语 , 英文术语**（用英文逗号分隔）")
            st.info(f"新增术语将归属到当前检测领域：**{st.session_state.term_best_domain}**")

            col_new_ch, col_new_en = st.columns([1, 1])
            with col_new_ch:
                new_chinese = st.text_input("中文术语", key="new_term_ch", placeholder="例如：临床试验")
            with col_new_en:
                new_english = st.text_input("英文术语", key="new_term_en", placeholder="例如：clinical trial")

            col_add, col_done, _ = st.columns([1, 1, 2])

            with col_add:
                if st.button("✅ 确认添加", type="primary", use_container_width=True):
                    ch = new_chinese.strip()
                    en = new_english.strip()
                    if ch and en:
                        # 校验前半为中文
                        if not any('一' <= c <= '鿿' for c in ch):
                            st.error("前半需为中文术语，格式：中文,英文")
                        else:
                            added = append_to_csv(ch, en, st.session_state.term_best_domain, csv_path)
                            if added:
                                st.success(f"✅ 已写入术语库：{ch} → {en}")
                            else:
                                st.info(f"术语已存在，跳过写入：{ch} → {en}")

                            # 记录到本次用户添加列表
                            if (ch, en) not in st.session_state.term_user_added:
                                st.session_state.term_user_added.append((ch, en))

                            # 刷新标注
                            text = st.session_state.term_input_text.strip()
                            detected2 = detect_domains(text)
                            if not detected2:
                                detected2 = detect_domains_by_terms(text, terminology)

                            if not detected2:
                                active_terms2: list[tuple[str, str]] = []
                                for t_list in terminology.values():
                                    active_terms2.extend(t_list)
                            else:
                                active_domains2 = [d for d, _ in detected2]
                                active_terms2 = []
                                for d in active_domains2:
                                    active_terms2.extend(terminology.get(d, []))

                            for ch_u, en_u in st.session_state.term_user_added:
                                if (ch_u, en_u) not in active_terms2:
                                    active_terms2.append((ch_u, en_u))

                            result2 = annotate_text(text, active_terms2)
                            st.session_state.term_annotated_segments = result2["annotated_segments"]
                            st.session_state.term_matched_csv = result2["matched_csv"]
                            st.session_state.term_matched_user = result2["matched_user"]
                            st.rerun()
                    else:
                        st.error("中文术语和英文术语均不能为空")

            with col_done:
                if st.button("✔️ 完成添加", use_container_width=True):
                    st.session_state.term_adding_mode = False
                    st.rerun()

            # 显示本轮已添加的术语
            if st.session_state.term_user_added:
                st.markdown("**本轮已添加：**")
                added_html = " ".join(
                    f'<span style="display:inline-block;background:#dc3545;color:#fff;padding:2px 8px;border-radius:10px;font-size:12px;margin:2px;">{ch} → {en}</span>'
                    for ch, en in st.session_state.term_user_added
                )
                st.markdown(added_html, unsafe_allow_html=True)

    else:
        if not st.session_state.term_annotated:
            st.info("👆 在上方输入文本，然后点击「标注术语」按钮开始分析。")

# ── 规则发现弹窗（全局）──────────────────────────────
if st.session_state.style_prompt:
    pd_data = st.session_state.style_prompt
    st.divider()
    st.warning(f"🔍 **发现稳定修订模式**\n\n以下相同修改已出现 **{pd_data['count']} 次**：\n\n**before_text**\n> {pd_data['original']}\n\n**after_text**\n> {pd_data['modified']}\n\n句对已自动存入 **Memory Base**。")
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("确认", type="primary", use_container_width=True): confirm_rule(pd_data["rule_id"]); st.session_state.style_prompt = None; st.success("已确认 🎉"); st.rerun()
    with col2:
        if st.button("忽略", use_container_width=True): ignore_rule(pd_data["rule_id"]); st.session_state.style_prompt = None; st.rerun()
    with col3:
        if st.button("以后再说", use_container_width=True): defer_rule(pd_data["rule_id"]); st.session_state.style_prompt = None; st.rerun()
